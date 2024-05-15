import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles.borders import Border, Side
import os 

def main():
    st.title("PFP Summary Table Creator")
    st.sidebar.title("Upload Files")

    mes = st.sidebar.number_input("Mes del archivo nuevo",value=None, placeholder="Mes")
    num_sets = st.sidebar.number_input("Numero de conjuntos de archivos", min_value=1, value=1)

    uploaded_files1 = []
    uploaded_files2 = []

    for _ in range(num_sets):
        uploaded_file1 = st.sidebar.file_uploader(f"Upload first file for set {_ + 1}", type=['xlsx'])
        uploaded_file2 = st.sidebar.file_uploader(f"Upload second file for set {_ + 1}", type=['xlsx'])
        uploaded_files1.append(uploaded_file1)
        uploaded_files2.append(uploaded_file2)

    for idx in range(num_sets):
        uploaded_file1 = uploaded_files1[idx]
        uploaded_file2 = uploaded_files2[idx]

        if uploaded_file1 and uploaded_file2:
            process_files(mes, uploaded_file1, uploaded_file2, idx)

def process_files(mes, uploaded_file1, uploaded_file2, idx):
    if uploaded_file1 and uploaded_file2 is not None:
        sourceWorkbook = load_workbook(uploaded_file1)
        workbook = load_workbook(uploaded_file2)

        sourceSheet = sourceWorkbook.active
        sheet = workbook.active

        #Locate target values
        def find_cell_with_value(sheet, column_index, target_value):
            for row_number, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index, values_only=True), start=1):
                for col_number, cell_value in enumerate(row, start=column_index):
                    if cell_value == target_value: 
                        sourceCellCoordinates = {
                            1 :[col_number+2, row_number],  #Resultado
                            2:[col_number, row_number+3],   #Table Origin cell
                            3:[col_number+1, row_number]    #Affected vol   
                        }
                        return sourceCellCoordinates
            return None  # Target value not found
        
        def copyRange(startCol, startRow, endCol, endRow, sheet):
            rangeSelected = []
            for i in range(startRow, endRow + 1,1):
                rowSelected = []
                for j in range(startCol, endCol+1,1):
                    rowSelected.append(sheet.cell(row = i, column = j ).value)
                rangeSelected.append(rowSelected)
            return rangeSelected
        
        def pasteRange(startCol, startRow, endCol, endRow, sheet, copiedData):
            countRow = 0
            for i in range(startRow, endRow+1, 1):
                countCol = 0
                for j in range(startCol, endCol+1, 1):
                    sheet.cell(row=i, column=j).value = copiedData[countRow][countCol]
                    countCol += 1
                countRow +=1

        sourceCellCoordinates = find_cell_with_value(sourceSheet, 3, 'TOTAL')
        newCellCoordinates = find_cell_with_value(sheet, 3, 'TOTAL')

        if mes == 4:
            objectiveApr = st.number_input("Ingrese el objetivo para mes Abril", value=None, placeholder="Abril")
            objectiveMay = st.number_input("Ingrese el objetivo para mes Mayo: ", value=None, placeholder="Mayo")
            objectiveJune = st.number_input("Ingrese el objetivo para mes Junio: ", value=None, placeholder="Junio")
            objectiveJuly = st.number_input("Ingrese el objetivo para mes Julio: ", value=None, placeholder="Julio")
            objectiveAug = st.number_input("Ingrese el objetivo para mes Agosto: ", value=None, placeholder="Agosto")
            objectiveSept = st.number_input("Ingrese el objetivo para mes Septiembre: ", value=None, placeholder="Septiembre")
            objectiveOct = st.number_input("Ingrese el objetivo para mes Octubre: ", value=None, placeholder="Octubre")
            objectiveNov = st.number_input("Ingrese el objetivo para mes Noviembre: ", value=None, placeholder="Noviembre")
            objectiveDec = st.number_input("Ingrese el objetivo para mes Diciembre: ", value=None, placeholder="Diciembre")
            objectiveJan = st.number_input("Ingrese el objetivo para mes Enero: ", value=None, placeholder="Enero")
            objectiveFeb = st.number_input("Ingrese el objetivo para mes Febrero: ", value=None, placeholder="Febrero")
            objectiveMar = st.number_input("Ingrese el objetivo para mes Marzo: ", value=None, placeholder="Marzo")
            realCREOY = st.number_input("Ingrese el Real CR EOY: ", value=None, placeholder="CR EOY")

            data = [
            ['FY23 Agreement and Result','','','','','',''],
            ["", "April", "May", "June", "July", "August", "September"],
            ["Objective", objectiveApr, objectiveMay, objectiveJune, objectiveJuly, objectiveAug, objectiveSept],
            ["Result", '', '', '', '', '', ''],
            ["CR", '', '', '', '', '', ''],
            ["Repaired per month", '', '', '', '', '', ''],
            ["Remaining to reach Objective", '', '', '', '', '', ''],
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "October", "November", "December", "January", "February", "March"],
            ["Objective", objectiveOct, objectiveNov, objectiveDec, objectiveJan, objectiveFeb, objectiveMar],
            ["Result", '', '', '', "", "", ""],
            ["CR", '', '', '', "", "", ""],
            ["Repaired per month", '', '', '', "", "", ""],
            ["Remaining to reach Objective", '', '', '', "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "Real CR FY23", ''],
            ["", "", "", "", "", "Real CR EOY FY23", realCREOY],
            ]

            pasteRange(newCellCoordinates[2][0], newCellCoordinates[2][1], (newCellCoordinates[2][0])+6, (newCellCoordinates[2][1])+17, sheet, data)

        else: 
            copiedData = copyRange(sourceCellCoordinates[2][0],sourceCellCoordinates[2][1], (sourceCellCoordinates[2][0]) +6, (sourceCellCoordinates[2][1]+17), sourceSheet)
            pasteRange(newCellCoordinates[2][0], newCellCoordinates[2][1], (newCellCoordinates[2][0])+6, (newCellCoordinates[2][1])+17, sheet, copiedData)

        col = newCellCoordinates[2][0]
        row = newCellCoordinates[2][1]

        month_mapping = {
            1:[col+4, row+9],
            2:[col+5, row+9],
            3:[col+6, row+9],
            4:[col+1, row+1],
            5:[col+2, row+1],
            6:[col+3, row+1],
            7:[col+4, row+1],
            8:[col+5, row+1],
            9:[col+6, row+1],
            10:[col+1, row+9],
            11:[col+2, row+9],
            12:[col+3, row+9],
        }

        affectedvol = sheet.cell(column=newCellCoordinates[3][0],row=newCellCoordinates[3][1]).value
        result = sheet.cell(column=newCellCoordinates[1][0],row=newCellCoordinates[1][1]).value
        objective = sheet.cell(column=month_mapping[mes][0],row=(month_mapping[mes][1])+1).value
        finalObjective = sheet.cell(column=month_mapping[3][0],row=(month_mapping[3][1])+1).value

        if mes == 1: 
            resultPreviouseMonth = sheet.cell(column=month_mapping[12][0],row=(month_mapping[12][1])+2).value
        elif mes == 4: 
            #Ask user to intoduce new values. 
            resultPreviouseMonth = sourceSheet.cell(column=month_mapping[3][0],row=(month_mapping[3][1])+2).value
        else:
            resultPreviouseMonth = sheet.cell(column=month_mapping[mes-1][0],row=(month_mapping[mes-1][1])+2).value

        cr = result/objective
        repairedPerMonth = result - resultPreviouseMonth
        remainingToReach = finalObjective - result
        realCR = result / affectedvol

        sheet.cell(column=month_mapping[mes][0], row=(month_mapping[mes][1])+2).value = result
        sheet.cell(column=month_mapping[mes][0], row=(month_mapping[mes][1])+3).value = cr
        sheet.cell(column=month_mapping[mes][0], row=(month_mapping[mes][1])+4).value = repairedPerMonth
        sheet.cell(column=month_mapping[mes][0], row=(month_mapping[mes][1])+5).value = remainingToReach
        sheet.cell(column=(newCellCoordinates[2][0])+6, row=(newCellCoordinates[2][1])+16).value = realCR

        number_format = '0.0%'
        sheet.cell(column=(newCellCoordinates[2][0])+6, row=(newCellCoordinates[2][1])+16).number_format = number_format
        sheet.cell(column=(newCellCoordinates[2][0])+6, row=(newCellCoordinates[2][1])+17).number_format = number_format
        for month in month_mapping:
            sheet.cell(column=month_mapping[month][0], row=(month_mapping[month][1])+3).number_format = number_format

        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        origin = newCellCoordinates[2][0]
        for i in range(1,7):
            sheet.cell(column=(newCellCoordinates[2][0])+i, row=(newCellCoordinates[2][1])+1).border = thin_border
        for i in range(0,7):
            for j in range(2,7):
                sheet.cell(column=(newCellCoordinates[2][0])+i, row=(newCellCoordinates[2][1])+j).border = thin_border
        for i in range(1,7):
            sheet.cell(column=(newCellCoordinates[2][0])+i, row=(newCellCoordinates[2][1])+9).border = thin_border
        for i in range(0,7):
            for j in range(10,15):
                sheet.cell(column=(newCellCoordinates[2][0])+i, row=(newCellCoordinates[2][1])+j).border = thin_border
        for i in range(5,7):
            for j in range(16,18):
                sheet.cell(column=(newCellCoordinates[2][0])+i, row=(newCellCoordinates[2][1])+j).border = thin_border




        output = BytesIO()
        output_filename = os.path.basename(uploaded_file2.name)
        workbook.save(output)
        output.seek(0)

        # Create a download button for the saved file
        st.download_button(f"Click here to download file - {idx}", output.getvalue(), file_name=output_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        
        #output_filename = os.path.basename(uploaded_file2.name)
        #workbook.save(output_filename)

        # Provide a download link for the output file
        #output_file = save_virtual_workbook(workbook)
        #st.download_button("Press to Download file:", output_file, output_filename)
        
    def save_virtual_workbook(workbook):
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()



def read_excel_file(uploaded_file):
    # Load the uploaded Excel file directly
    wb = load_workbook(uploaded_file)
    # Assuming there's only one sheet, you can access it like this:
    sheet = wb.active
    # Read the data from the sheet and return it
    excel_data = []
    for row in sheet.iter_rows(values_only=True):
        excel_data.append(row)
    return excel_data

if __name__ == "__main__":
    main()
